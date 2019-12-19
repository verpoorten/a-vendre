from openpyxl import Workbook
from openpyxl.utils import get_column_letter
MAX_COL_WIDTH = 50


def adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
                continue
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2

        if adjusted_width > MAX_COL_WIDTH:
            adjusted_width = MAX_COL_WIDTH
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
        else:
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

